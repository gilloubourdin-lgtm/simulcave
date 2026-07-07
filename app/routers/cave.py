# app/routers/cave.py

import csv
import io
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font

from fastapi import APIRouter, Depends, Form, HTTPException, Request, Header
from fastapi.responses import RedirectResponse, Response, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Cave, Wall, Zone, User, RenovationScenarioDB, ZoneMonthlyTarget
from app.routers.auth import require_user
from app.services.simulation import simulate_cave
from app.services.materials import get_material_properties
from app.services.renovation import (
    generate_renovation_scenarios,
    calculate_scenario,
)
from app.services.pdf import generate_cave_report_pdf
from app.services.geocoding import geocode_address
from app.services.nrcave_export import build_nrcave_payload

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


def render_template(
    request: Request,
    template_name: str,
    context: dict | None = None,
):
    if context is None:
        context = {}

    return templates.TemplateResponse(
        request=request,
        name=template_name,
        context=context,
    )


def get_user_cave(
    db: Session,
    cave_id: int,
    current_user: User,
) -> Cave:
    cave = db.query(Cave).filter(
        Cave.id == cave_id,
        Cave.user_id == current_user.id,
    ).first()

    if not cave:
        raise HTTPException(status_code=404, detail="Cave introuvable.")

    return cave


@router.get("/")
def home():
    return RedirectResponse(url="/login", status_code=303)

@router.get("/dashboard")
def dashboard(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    caves = db.query(Cave).filter(
        Cave.user_id == current_user.id,
    ).all()

    total_energy = 0
    total_cost = 0
    total_co2 = 0
    cave_count = len(caves)

    results = []

    for cave in caves:
        result = simulate_cave(cave)

        total_energy += result.total_energy_kwh
        total_cost += result.annual_cost_chf
        total_co2 += result.annual_co2_tons

        results.append({
            "cave": cave,
            "result": result,
        })

    best_cave = None
    worst_cave = None
    average_cost = 0
    average_energy_per_m3 = 0

    if results:
        best_cave = min(
            results,
            key=lambda x: x["result"].total_energy_kwh,
        )

        worst_cave = max(
            results,
            key=lambda x: x["result"].total_energy_kwh,
        )

        average_cost = total_cost / cave_count

        total_volume = sum(
            item["cave"].length_m * item["cave"].width_m * item["cave"].height_m
            for item in results
        )

        if total_volume > 0:
            average_energy_per_m3 = total_energy / total_volume

    results_sorted = sorted(
        results,
        key=lambda x: x["result"].total_energy_kwh,
        reverse=True,
    )

    return render_template(
        request,
        "dashboard.html",
        {
            "cave_count": cave_count,
            "total_energy": round(total_energy, 1),
            "total_cost": round(total_cost, 0),
            "total_co2": round(total_co2, 2),
            "results": results_sorted,
            "best_cave": best_cave,
            "worst_cave": worst_cave,
            "average_cost": round(average_cost, 1),
            "average_energy_per_m3": round(average_energy_per_m3, 2),
        },
    )

@router.get("/caves")
def caves_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    caves = db.query(Cave).filter(
        Cave.user_id == current_user.id,
    ).all()

    cave_summaries = []

    for cave in caves:
        result = simulate_cave(cave)

        cave_summaries.append({
            "cave": cave,
            "result": result,
            "volume_m3": round(cave.length_m * cave.width_m * cave.height_m, 1),
        })

    return render_template(
        request,
        "caves_list.html",
        {"cave_summaries": cave_summaries},
    )


@router.get("/caves/new")
def cave_form(
    request: Request,
    current_user: User = Depends(require_user),
):
    return render_template(
        request,
        "cave_form.html",
    )


@router.post("/caves")
def create_cave(
    name: str = Form(...),
    region: str = Form(...),
    address: str = Form(""),
    latitude: float | None = Form(None),
    longitude: float | None = Form(None),
    use_dynamic_weather: str | None = Form(None),
    altitude_m: float = Form(...),
    length_m: float = Form(...),
    width_m: float = Form(...),
    height_m: float = Form(...),
    buried_factor: float = Form(...),
    wall_n_material: str = Form(...),
    wall_n_u: float = Form(...),
    wall_s_material: str = Form(...),
    wall_s_u: float = Form(...),
    wall_e_material: str = Form(...),
    wall_e_u: float = Form(...),
    wall_w_material: str = Form(...),
    wall_w_u: float = Form(...),
    roof_material: str = Form(...),
    roof_u: float = Form(...),
    floor_material: str = Form(...),
    floor_u: float = Form(...),
    zone_count: int = Form(...),
    energy_source: str = Form(...),
    energy_price_chf_per_kwh: float = Form(...),
    co2_factor_kg_per_kwh: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    zone_count = max(int(zone_count), 1)

    dynamic_weather_enabled = use_dynamic_weather == "true"

    if dynamic_weather_enabled and (latitude is None or longitude is None):
        place = geocode_address(address)
        if place:
            latitude = place["latitude"]
            longitude = place["longitude"]

    cave = Cave(
        name=name,
        user_id=current_user.id,
        region=region,
        altitude_m=altitude_m,
        length_m=length_m,
        width_m=width_m,
        height_m=height_m,
        buried_factor=buried_factor,
        energy_source=energy_source,
        energy_price_chf_per_kwh=energy_price_chf_per_kwh,
        co2_factor_kg_per_kwh=co2_factor_kg_per_kwh,
        address=address,
        latitude=latitude,
        longitude=longitude,
        use_dynamic_weather=dynamic_weather_enabled,
        ventilation_enabled=True,
        ventilation_rate_ach=0.10,
    )

    db.add(cave)
    db.flush()

    wall_height_area_long = length_m * height_m
    wall_height_area_short = width_m * height_m
    roof_floor_area = length_m * width_m

    wall_inputs = [
        ("Mur Nord", "N", wall_n_material, wall_height_area_long, wall_n_u),
        ("Mur Sud", "S", wall_s_material, wall_height_area_long, wall_s_u),
        ("Mur Est", "E", wall_e_material, wall_height_area_short, wall_e_u),
        ("Mur Ouest", "O", wall_w_material, wall_height_area_short, wall_w_u),
        ("Toiture", "H", roof_material, roof_floor_area, roof_u),
        ("Sol", "B", floor_material, roof_floor_area, floor_u),
    ]

    walls = []

    for wall_name, orientation, material, area, u_value in wall_inputs:
        props = get_material_properties(material)

        walls.append(
            Wall(
                cave_id=cave.id,
                name=wall_name,
                orientation=orientation,
                material=material,
                area_m2=area,
                u_value=u_value,
                thickness_m=props["default_thickness_m"],
                inertia_factor=props["inertia_factor"],
            )
        )

    db.add_all(walls)

    default_monthly_profile = {
        1: (10, 75, "FML / stabilisation"),
        2: (10, 75, "Stabilisation"),
        3: (12, 75, "Stockage"),
        4: (12, 75, "Stockage"),
        5: (13, 75, "Stockage"),
        6: (14, 75, "Été"),
        7: (16, 75, "Été"),
        8: (16, 75, "Été"),
        9: (13, 75, "Refroidissement FA"),
        10: (10, 75, "FA / FML"),
        11: (8, 75, "Refroidissement FA"),
        12: (8, 75, "FML"),
    }

    total_volume = length_m * width_m * height_m
    default_zone_volume = total_volume / zone_count
    zone_length = length_m / zone_count

    zones = []

    for i in range(zone_count):
        zone = Zone(
            cave_id=cave.id,
            name=f"Zone {i + 1}",
            volume_m3=default_zone_volume,
            target_temp_winter_c=12,
            target_temp_summer_c=16,
            target_humidity_percent=75,
            process_cooling_kwh=0,
            process_heating_kwh=0,
            x_m=i * zone_length,
            y_m=0,
            width_m=width_m,
            length_m=zone_length,
            process_heating_start_month=1,
            process_heating_end_month=12,
            process_cooling_start_month=1,
            process_cooling_end_month=12,
        )

        db.add(zone)
        db.flush()

        for month, values in default_monthly_profile.items():
            temp, humidity, phase = values

            db.add(
                ZoneMonthlyTarget(
                    zone_id=zone.id,
                    month=month,
                    target_temp_c=temp,
                    target_humidity_percent=humidity,
                    phase=phase,
                )
            )

        zones.append(zone)

    print(f"Cave créée avec {len(zones)} zone(s)")

    db.commit()

    return RedirectResponse(
        url=f"/caves/{cave.id}",
        status_code=303,
    )

@router.get("/caves/{cave_id}/edit")
def edit_cave_form(
    request: Request,
    cave_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    return templates.TemplateResponse(
        request=request,
        name="cave_edit.html",
        context={
            "cave": cave,
        },
    )


@router.post("/caves/{cave_id}/edit")
def edit_cave_submit(
    request: Request,
    cave_id: int,
    name: str = Form(...),
    region: str = Form(...),
    energy_price_chf_per_kwh: float = Form(...),
    co2_factor_kg_per_kwh: float = Form(...),
    ventilation_rate_ach: float = Form(0.10),
    ventilation_enabled: str | None = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    cave.name = name
    cave.region = region
    cave.energy_price_chf_per_kwh = energy_price_chf_per_kwh
    cave.co2_factor_kg_per_kwh = co2_factor_kg_per_kwh
    cave.ventilation_rate_ach = ventilation_rate_ach
    cave.ventilation_enabled = ventilation_enabled == "on"

    db.commit()

    return RedirectResponse(
        url=f"/caves/{cave.id}",
        status_code=303,
    )

@router.get("/caves/compare")
def compare_caves(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    caves = db.query(Cave).filter(
        Cave.user_id == current_user.id
    ).all()

    comparison = []

    for cave in caves:
        result = simulate_cave(cave)

        comparison.append({
            "cave": cave,
            "result": result,
        })

    return render_template(
        request,
        "compare_caves.html",
        {
            "comparison": comparison,
        },
    )

@router.get("/caves/{cave_id}")
def cave_detail(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)
    result = simulate_cave(cave)

    return render_template(
        request,
        "cave_detail.html",
        {
            "cave": cave,
            "result": result,
        },
    )

@router.get("/zones/{zone_id}/update")
def update_zone_form(
    zone_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    zone = db.query(Zone).join(Cave).filter(
        Zone.id == zone_id,
        Cave.user_id == current_user.id,
    ).first()

    if not zone:
        raise HTTPException(
            status_code=404,
            detail="Zone introuvable.",
        )

    return render_template(
        request,
        "zone_edit.html",
        {
            "zone": zone,
        },
    )

@router.post("/zones/{zone_id}/update")
def update_zone(
    zone_id: int,
    name: str = Form(...),
    volume_m3: float = Form(...),
    x_m: float = Form(0),
    y_m: float = Form(0),
    width_m: float = Form(1),
    length_m: float = Form(1),
    target_temp_winter_c: float = Form(...),
    target_temp_summer_c: float = Form(...),
    target_humidity_percent: float = Form(...),
    process_cooling_kwh: float = Form(0),
    process_heating_kwh: float = Form(0),
    process_heating_start_month: int = Form(1),
    process_heating_end_month: int = Form(12),
    process_cooling_start_month: int = Form(1),
    process_cooling_end_month: int = Form(12),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    level_index: int = Form(0),
    level_name: str = Form("Rez"),
    floor_depth_m: float = Form(0),
):
    zone = db.query(Zone).join(Cave).filter(
        Zone.id == zone_id,
        Cave.user_id == current_user.id,
    ).first()

    if not zone:
        raise HTTPException(status_code=404, detail="Zone introuvable.")

    zone.name = name
    zone.volume_m3 = volume_m3
    zone.x_m = x_m
    zone.y_m = y_m
    zone.width_m = width_m
    zone.length_m = length_m
    zone.target_temp_winter_c = target_temp_winter_c
    zone.target_temp_summer_c = target_temp_summer_c
    zone.target_humidity_percent = target_humidity_percent
    zone.process_cooling_kwh = process_cooling_kwh
    zone.process_heating_kwh = process_heating_kwh
    zone.process_heating_start_month = process_heating_start_month
    zone.process_heating_end_month = process_heating_end_month
    zone.process_cooling_start_month = process_cooling_start_month
    zone.process_cooling_end_month = process_cooling_end_month
    zone.level_index = level_index
    zone.level_name = level_name
    zone.floor_depth_m = floor_depth_m

    cave_id = zone.cave_id
    db.commit()

    return RedirectResponse(url=f"/caves/{cave_id}", status_code=303)


@router.get("/caves/{cave_id}/simulate")
def simulate(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)
    result = simulate_cave(cave)

    return render_template(
        request,
        "simulation_result.html",
        {
            "cave": cave,
            "result": result,
        },
    )


@router.get("/caves/{cave_id}/renovation")
def renovation(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    scenarios = generate_renovation_scenarios(cave)

    saved_scenarios_results = []

    for saved in cave.renovation_scenarios:
        reductions = []

        if saved.roof_reduction_percent > 0:
            reductions.append(saved.roof_reduction_percent)

        if saved.walls_reduction_percent > 0:
            reductions.append(saved.walls_reduction_percent)

        if saved.floor_reduction_percent > 0:
            reductions.append(saved.floor_reduction_percent)

        average_reduction_percent = sum(reductions) / len(reductions) if reductions else 0
        reduction_factor = 1 - (average_reduction_percent / 100)

        def wall_filter(wall, saved=saved):
            if wall.name == "Toiture" and saved.roof_reduction_percent > 0:
                return True

            if wall.orientation in ["N", "S", "E", "O"] and saved.walls_reduction_percent > 0:
                return True

            if wall.name == "Sol" and saved.floor_reduction_percent > 0:
                return True

            return False

        calculated = calculate_scenario(
            cave=cave,
            name=saved.name,
            description=(
                f"Toiture: -{saved.roof_reduction_percent} %, "
                f"murs: -{saved.walls_reduction_percent} %, "
                f"sol: -{saved.floor_reduction_percent} %."
            ),
            investment_chf=saved.investment_chf,
            wall_filter=wall_filter,
            reduction_factor=reduction_factor,
        )

        saved_scenarios_results.append({
            "saved": saved,
            "calculated": calculated,
        })

    return render_template(
        request,
        "renovation_result.html",
        {
            "cave": cave,
            "scenarios": scenarios,
            "saved_scenarios_results": saved_scenarios_results,
        },
    )


@router.get("/caves/{cave_id}/renovation/custom")
def renovation_custom_form(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    return render_template(
        request,
        "renovation_custom_form.html",
        {"cave": cave},
    )

@router.get("/caves/{cave_id}/renovation/compare")
def renovation_compare(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    baseline = simulate_cave(cave)
    scenarios = generate_renovation_scenarios(cave)

    comparisons = []

    for scenario in scenarios:
        energy_saving_percent = 0

        if baseline.total_energy_kwh > 0:
            energy_saving_percent = (
                scenario.energy_saved_kwh
                / baseline.total_energy_kwh
                * 100
            )

        comparisons.append({
            "scenario": scenario,
            "baseline_energy_kwh": baseline.total_energy_kwh,
            "renovated_energy_kwh": baseline.total_energy_kwh - scenario.energy_saved_kwh,
            "energy_saving_percent": round(energy_saving_percent, 1),
            "baseline_cost_chf": baseline.annual_cost_chf,
            "renovated_cost_chf": baseline.annual_cost_chf - scenario.money_saved_chf,
            "baseline_co2_tons": baseline.annual_co2_tons,
            "co2_saving_tons": round(scenario.co2_saved_kg / 1000, 2),
        })

    return render_template(
        request,
        "renovation_compare.html",
        {
            "cave": cave,
            "baseline": baseline,
            "comparisons": comparisons,
        },
    )


@router.post("/caves/{cave_id}/renovation/custom")
def renovation_custom_result(
    cave_id: int,
    request: Request,
    scenario_name: str = Form(...),
    investment_chf: float = Form(...),
    roof_reduction_percent: float = Form(...),
    walls_reduction_percent: float = Form(...),
    floor_reduction_percent: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    def wall_filter(wall):
        if wall.name == "Toiture" and roof_reduction_percent > 0:
            return True

        if wall.orientation in ["N", "S", "E", "O"] and walls_reduction_percent > 0:
            return True

        if wall.name == "Sol" and floor_reduction_percent > 0:
            return True

        return False

    reductions = []

    if roof_reduction_percent > 0:
        reductions.append(roof_reduction_percent)

    if walls_reduction_percent > 0:
        reductions.append(walls_reduction_percent)

    if floor_reduction_percent > 0:
        reductions.append(floor_reduction_percent)

    average_reduction_percent = sum(reductions) / len(reductions) if reductions else 0
    reduction_factor = 1 - (average_reduction_percent / 100)

    scenario = calculate_scenario(
        cave=cave,
        name=scenario_name,
        description=(
            f"Toiture: -{roof_reduction_percent} %, "
            f"murs: -{walls_reduction_percent} %, "
            f"sol: -{floor_reduction_percent} % sur les valeurs U."
        ),
        investment_chf=investment_chf,
        wall_filter=wall_filter,
        reduction_factor=reduction_factor,
    )

    return render_template(
        request,
        "renovation_custom_result.html",
        {
            "cave": cave,
            "scenario": scenario,
            "roof_reduction_percent": roof_reduction_percent,
            "walls_reduction_percent": walls_reduction_percent,
            "floor_reduction_percent": floor_reduction_percent,
        },
    )


@router.get("/caves/{cave_id}/report/pdf")
def cave_report_pdf(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)
    pdf_bytes = generate_cave_report_pdf(request, cave)

    filename = f"rapport_simulcave_{cave.id}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )

@router.get("/caves/{cave_id}/plan")
def cave_plan(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    result = simulate_cave(cave)

    return render_template(
        request,
        "cave_plan.html",
        {
            "cave": cave,
            "result": result,
        },
    )

@router.get("/caves/{cave_id}/parameters")
def cave_parameters(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    result = simulate_cave(cave)

    return render_template(
        request,
        "cave_parameters.html",
        {
            "cave": cave,
            "result": result,
        },
    )

@router.post("/walls/{wall_id}/update")
def update_wall(
    wall_id: int,
    material: str = Form(...),
    area_m2: float = Form(...),
    u_value: float = Form(...),
    thickness_m: float = Form(...),
    inertia_factor: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    wall = db.query(Wall).join(Cave).filter(
        Wall.id == wall_id,
        Cave.user_id == current_user.id,
    ).first()

    if not wall:
        raise HTTPException(status_code=404, detail="Paroi introuvable.")

    wall.material = material
    wall.area_m2 = area_m2
    wall.u_value = u_value
    wall.thickness_m = thickness_m
    wall.inertia_factor = inertia_factor

    cave_id = wall.cave_id

    db.commit()

    return RedirectResponse(
        url=f"/caves/{cave_id}",
        status_code=303,
    )

@router.post("/caves/{cave_id}/renovation/custom/save")
def save_custom_renovation_scenario(
    cave_id: int,
    scenario_name: str = Form(...),
    investment_chf: float = Form(...),
    roof_reduction_percent: float = Form(...),
    walls_reduction_percent: float = Form(...),
    floor_reduction_percent: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    scenario = RenovationScenarioDB(
        cave_id=cave.id,
        name=scenario_name,
        investment_chf=investment_chf,
        roof_reduction_percent=roof_reduction_percent,
        walls_reduction_percent=walls_reduction_percent,
        floor_reduction_percent=floor_reduction_percent,
    )

    db.add(scenario)
    db.commit()

    return RedirectResponse(
        url=f"/caves/{cave.id}/renovation",
        status_code=303,
    )

@router.get("/caves/{cave_id}/export/csv")
def export_cave_csv(
    cave_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)
    result = simulate_cave(cave)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    writer.writerow(["SimulCave - Export simulation"])
    writer.writerow(["Cave", cave.name])
    writer.writerow(["Région", cave.region])
    writer.writerow(["Adresse", cave.address or ""])
    writer.writerow(["Latitude", cave.latitude or ""])
    writer.writerow(["Longitude", cave.longitude or ""])
    writer.writerow(["Source météo", result.weather_source])
    writer.writerow(["Température sol [°C]", result.soil_temperature_c])
    writer.writerow([])

    writer.writerow([
        "Mois",
        "Température extérieure [°C]",
        "Température effective [°C]",
        "Chaud [kWh]",
        "Froid [kWh]",
    ])

    for month in result.monthly_results:
        writer.writerow([
            month.month,
            month.outdoor_temp_c,
            month.effective_temp_c,
            month.heating_kwh,
            month.cooling_kwh,
        ])

    writer.writerow([])
    writer.writerow([
        "Paroi",
        "Orientation",
        "Matériau",
        "Chaud [kWh/an]",
        "Froid [kWh/an]",
        "Total [kWh/an]",
    ])

    for wall in result.wall_results:
        writer.writerow([
            wall.wall_name,
            wall.orientation,
            wall.material,
            wall.heating_kwh,
            wall.cooling_kwh,
            wall.total_kwh,
        ])

    csv_content = output.getvalue()

    return Response(
        content=csv_content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="simulcave_{cave.id}.csv"'
        },
    )

@router.get("/caves/{cave_id}/export/xlsx")
def export_cave_xlsx(
    cave_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)
    result = simulate_cave(cave)

    wb = Workbook()

    ws = wb.active
    ws.title = "Synthèse"

    rows = [
        ["Cave", cave.name],
        ["Région", cave.region],
        ["Adresse", cave.address or ""],
        ["Latitude", cave.latitude or ""],
        ["Longitude", cave.longitude or ""],
        ["Source météo", result.weather_source],
        ["Température sol [°C]", result.soil_temperature_c],
        ["Énergie totale [kWh/an]", result.total_energy_kwh],
        ["Coût annuel [CHF/an]", result.annual_cost_chf],
        ["CO₂ [t/an]", result.annual_co2_tons],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True)

    ws_month = wb.create_sheet("Mensuel")
    ws_month.append([
        "Mois",
        "Température extérieure [°C]",
        "Température effective [°C]",
        "Chaud [kWh]",
        "Froid [kWh]",
    ])

    for month in result.monthly_results:
        ws_month.append([
            month.month,
            month.outdoor_temp_c,
            month.effective_temp_c,
            month.heating_kwh,
            month.cooling_kwh,
        ])

    ws_walls = wb.create_sheet("Parois")
    ws_walls.append([
        "Paroi",
        "Orientation",
        "Matériau",
        "Surface [m²]",
        "U [W/m²K]",
        "Épaisseur [m]",
        "Inertie",
        "Chaud [kWh/an]",
        "Froid [kWh/an]",
        "Total [kWh/an]",
    ])

    wall_results_by_name = {
        wall.wall_name: wall for wall in result.wall_results
    }

    for wall in cave.walls:
        wall_result = wall_results_by_name.get(wall.name)

        ws_walls.append([
            wall.name,
            wall.orientation,
            wall.material,
            wall.area_m2,
            wall.u_value,
            wall.thickness_m,
            wall.inertia_factor,
            wall_result.heating_kwh if wall_result else "",
            wall_result.cooling_kwh if wall_result else "",
            wall_result.total_kwh if wall_result else "",
        ])

    ws_zones = wb.create_sheet("Zones")
    ws_zones.append([
        "Zone",
        "Volume [m³]",
        "X [m]",
        "Y [m]",
        "Longueur [m]",
        "Largeur [m]",
        "Temp. hiver [°C]",
        "Temp. été [°C]",
        "Humidité [%]",
        "Chaud process [kWh/an]",
        "Période chaud",
        "Froid process [kWh/an]",
        "Période froid",
    ])

    for zone in cave.zones:
        ws_zones.append([
            zone.name,
            zone.volume_m3,
            zone.x_m,
            zone.y_m,
            zone.length_m,
            zone.width_m,
            zone.target_temp_winter_c,
            zone.target_temp_summer_c,
            zone.target_humidity_percent,
            zone.process_heating_kwh,
            f"{zone.process_heating_start_month}-{zone.process_heating_end_month}",
            zone.process_cooling_kwh,
            f"{zone.process_cooling_start_month}-{zone.process_cooling_end_month}",
        ])

    ws_scenarios = wb.create_sheet("Scénarios rénovation")

    ws_scenarios.append([
        "Nom",
        "Investissement [CHF]",
        "Réduction toiture [%]",
        "Réduction murs [%]",
        "Réduction sol [%]",
        "Économie énergie [kWh/an]",
        "Économie [CHF/an]",
        "CO₂ évité [kg/an]",
        "ROI [ans]",
    ])

    for saved in cave.renovation_scenarios:
        reductions = []

        if saved.roof_reduction_percent > 0:
            reductions.append(saved.roof_reduction_percent)

        if saved.walls_reduction_percent > 0:
            reductions.append(saved.walls_reduction_percent)

        if saved.floor_reduction_percent > 0:
            reductions.append(saved.floor_reduction_percent)

        average_reduction_percent = sum(reductions) / len(reductions) if reductions else 0
        reduction_factor = 1 - (average_reduction_percent / 100)

        def wall_filter(wall, saved=saved):
            if wall.name == "Toiture" and saved.roof_reduction_percent > 0:
                return True
            if wall.orientation in ["N", "S", "E", "O"] and saved.walls_reduction_percent > 0:
                return True
            if wall.name == "Sol" and saved.floor_reduction_percent > 0:
                return True
            return False

        calculated = calculate_scenario(
            cave=cave,
            name=saved.name,
            description="Scénario sauvegardé",
            investment_chf=saved.investment_chf,
            wall_filter=wall_filter,
            reduction_factor=reduction_factor,
        )

        ws_scenarios.append([
            saved.name,
            saved.investment_chf,
            saved.roof_reduction_percent,
            saved.walls_reduction_percent,
            saved.floor_reduction_percent,
            calculated.energy_saved_kwh,
            calculated.money_saved_chf,
            calculated.co2_saved_kg,
            calculated.payback_years if calculated.payback_years else "Non rentable",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="simulcave_{cave.id}.xlsx"'
        },
    )

@router.post("/renovation-scenarios/{scenario_id}/update")
def update_saved_renovation_scenario(
    scenario_id: int,
    name: str = Form(...),
    investment_chf: float = Form(...),
    roof_reduction_percent: float = Form(...),
    walls_reduction_percent: float = Form(...),
    floor_reduction_percent: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    scenario = db.query(RenovationScenarioDB).join(Cave).filter(
        RenovationScenarioDB.id == scenario_id,
        Cave.user_id == current_user.id,
    ).first()

    if not scenario:
        raise HTTPException(
            status_code=404,
            detail="Scénario introuvable.",
        )

    scenario.name = name
    scenario.investment_chf = investment_chf
    scenario.roof_reduction_percent = roof_reduction_percent
    scenario.walls_reduction_percent = walls_reduction_percent
    scenario.floor_reduction_percent = floor_reduction_percent

    db.commit()

    return RedirectResponse(
        url=f"/caves/{scenario.cave_id}/renovation",
        status_code=303,
    )

@router.post("/renovation-scenarios/{scenario_id}/duplicate")
def duplicate_saved_renovation_scenario(
    scenario_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    scenario = db.query(RenovationScenarioDB).join(Cave).filter(
        RenovationScenarioDB.id == scenario_id,
        Cave.user_id == current_user.id,
    ).first()

    if not scenario:
        raise HTTPException(
            status_code=404,
            detail="Scénario introuvable.",
        )

    duplicated = RenovationScenarioDB(
        cave_id=scenario.cave_id,
        name=f"{scenario.name} - copie",
        investment_chf=scenario.investment_chf,
        roof_reduction_percent=scenario.roof_reduction_percent,
        walls_reduction_percent=scenario.walls_reduction_percent,
        floor_reduction_percent=scenario.floor_reduction_percent,
    )

    db.add(duplicated)
    db.commit()

    return RedirectResponse(
        url=f"/caves/{scenario.cave_id}/renovation",
        status_code=303,
    )

@router.post("/renovation-scenarios/{scenario_id}/delete")
def delete_saved_renovation_scenario(
    scenario_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    scenario = db.query(RenovationScenarioDB).join(Cave).filter(
        RenovationScenarioDB.id == scenario_id,
        Cave.user_id == current_user.id,
    ).first()

    if not scenario:
        raise HTTPException(status_code=404, detail="Scénario introuvable.")

    cave_id = scenario.cave_id

    db.delete(scenario)
    db.commit()

    return RedirectResponse(
        url=f"/caves/{cave_id}/renovation",
        status_code=303,
    )

@router.post("/caves/{cave_id}/delete")
def delete_cave(
    cave_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    db.delete(cave)
    db.commit()

    return RedirectResponse(url="/caves", status_code=303)

@router.get("/export/nrcave/{cave_id}")
def export_for_nrcave(
    cave_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    result = simulate_cave(cave)

    thermal_needs = []
    cooling_needs = []
    temperature_monthly = []
    humidity_monthly = []
    humidity_risk_index_monthly = []

    for month in result.monthly_results:
        thermal_needs.append(float(month.heating_kwh or 0.0))
        cooling_needs.append(float(month.cooling_kwh or 0.0))
        temperature_monthly.append(float(month.effective_temp_c or 0.0))

        humidity_monthly.append(
            float(month.relative_humidity_percent or 75.0)
        )

        humidity_risk_index_monthly.append(
            float(month.humidity_risk_index or 0.0)
        )

    total_thermal_kwh = sum(thermal_needs)

    recommended_hvac_power_kw = 0.0
    if total_thermal_kwh > 0:
        recommended_hvac_power_kw = round(total_thermal_kwh / 1800.0, 1)

    summer_cooling = sum(cooling_needs[5:8])

    passive_cooling_ratio = 0.0
    if summer_cooling > 0:
        passive_cooling_ratio = max(
            0.0,
            min(1.0, 1.0 - (summer_cooling / 50000.0)),
        )

    if temperature_monthly:
        thermal_amplitude = max(temperature_monthly) - min(temperature_monthly)
    else:
        thermal_amplitude = 0.0

    thermal_stability_index = max(
        0.0,
        min(100.0, 100.0 - thermal_amplitude * 6.0),
    )

    simulation_result = {
        "thermal_needs_kwh_monthly": thermal_needs,
        "cooling_needs_kwh_monthly": cooling_needs,
        "humidity_risk_index_monthly": humidity_risk_index_monthly,
        "recommended_hvac_power_kw": recommended_hvac_power_kw,
        "temperature_monthly": temperature_monthly,
        "humidity_monthly": humidity_monthly,
        "thermal_stability_index": round(thermal_stability_index, 1),
        "passive_cooling_ratio": round(passive_cooling_ratio, 3),
    }

    payload = build_nrcave_payload(simulation_result)

    return Response(
        content=json.dumps(payload, indent=2, ensure_ascii=False),
        media_type="application/json",
    )

@router.get("/caves/{cave_id}/monthly-targets")
def monthly_targets_form(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)

    months = [
        (1, "Jan"),
        (2, "Fév"),
        (3, "Mar"),
        (4, "Avr"),
        (5, "Mai"),
        (6, "Juin"),
        (7, "Juil"),
        (8, "Août"),
        (9, "Sep"),
        (10, "Oct"),
        (11, "Nov"),
        (12, "Déc"),
    ]

    for zone in cave.zones:
        existing_months = {
            target.month for target in zone.monthly_targets
        }

        for month_number, _ in months:
            if month_number not in existing_months:
                fallback_temp = (
                    zone.target_temp_winter_c
                    if month_number in [1, 2, 3, 11, 12]
                    else zone.target_temp_summer_c
                )

                db.add(
                    ZoneMonthlyTarget(
                        zone_id=zone.id,
                        month=month_number,
                        target_temp_c=fallback_temp,
                        target_humidity_percent=zone.target_humidity_percent or 75,
                        phase="standard",
                    )
                )

    db.commit()
    db.refresh(cave)

    return render_template(
        request,
        "monthly_targets.html",
        {
            "cave": cave,
            "months": months,
        },
    )


@router.post("/caves/{cave_id}/monthly-targets")
async def monthly_targets_submit(
    cave_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    cave = get_user_cave(db, cave_id, current_user)
    form = await request.form()

    for zone in cave.zones:
        for month in range(1, 13):
            temp_key = f"temp_{zone.id}_{month}"
            humidity_key = f"humidity_{zone.id}_{month}"
            phase_key = f"phase_{zone.id}_{month}"

            target = db.query(ZoneMonthlyTarget).filter(
                ZoneMonthlyTarget.zone_id == zone.id,
                ZoneMonthlyTarget.month == month,
            ).first()

            if not target:
                target = ZoneMonthlyTarget(
                    zone_id=zone.id,
                    month=month,
                )
                db.add(target)

            target.target_temp_c = float(str(form.get(temp_key, 12)).replace(",", "."))
            target.target_humidity_percent = float(str(form.get(humidity_key, 75)).replace(",", "."))
            target.phase = str(form.get(phase_key, "standard"))

    db.commit()

    return RedirectResponse(
        url=f"/caves/{cave.id}",
        status_code=303,
    )

SIMULCAVE_API_TOKEN = os.getenv("SIMULCAVE_API_TOKEN", "")


@router.get("/api/caves/{cave_id}/building-summary")
def api_building_summary(
    cave_id: int,
    x_simulcave_token: str | None = Header(None),
    db: Session = Depends(get_db),
):
    if SIMULCAVE_API_TOKEN:
        if x_simulcave_token != SIMULCAVE_API_TOKEN:
            raise HTTPException(status_code=401, detail="Invalid API token")

    cave = db.query(Cave).filter(Cave.id == cave_id).first()

    if not cave:
        raise HTTPException(status_code=404, detail="Cave introuvable.")

    result = simulate_cave(cave)

    return {
        "source": "simulcave",
        "cave_id": cave.id,
        "cave_name": cave.name,
        "building_only": True,

        "climate_score": result.climate_score,
        "climate_label": result.climate_label,
        "energy_class": result.energy_class,
        "energy_intensity_kwh_m3": result.energy_intensity_kwh_m3,

        "annual": {
            "heating_kwh": result.total_heating_kwh,
            "cooling_kwh": result.total_cooling_kwh,
            "energy_kwh": result.total_energy_kwh,
            "cost_chf": result.annual_cost_chf,
            "co2_kg": result.annual_co2_kg,
            "co2_tons": result.annual_co2_tons,
        },

        "monthly": [
            {
                "month": month.month,
                "heating_kwh": month.heating_kwh,
                "cooling_kwh": month.cooling_kwh,
                "effective_temp_c": month.effective_temp_c,
                "relative_humidity_percent": month.relative_humidity_percent,
                "humidity_risk_index": getattr(month, "humidity_risk_index", 0),
                "condensation_risk": month.condensation_risk,
            }
            for month in result.monthly_results
        ],
    }

NRCAVE_API_TOKEN = os.getenv("NRCAVE_API_TOKEN", "")


@router.post("/api/nrcave/import")
def import_from_nrcave(
    payload: dict,
    x_nrcave_token: str | None = Header(None),
    db: Session = Depends(get_db),
):
    if NRCAVE_API_TOKEN:
        if x_nrcave_token != NRCAVE_API_TOKEN:
            raise HTTPException(status_code=401, detail="Invalid NRCave token")

    data = payload.get("cave", payload)

    name = data.get("name") or payload.get("name") or "Cave importée NRCave"
    region = data.get("region") or data.get("canton") or "Vaud"

    cave = Cave(
        name=name,
        region=region,
        address=data.get("address"),
        altitude_m=float(data.get("altitude_m") or 400),
        length_m=float(data.get("length_m") or 20),
        width_m=float(data.get("width_m") or 10),
        height_m=float(data.get("height_m") or 4),
        buried_factor=float(data.get("buried_factor") or 0.5),
        energy_source=data.get("energy_source") or "electricity",
        energy_price_chf_per_kwh=float(data.get("energy_price_chf_per_kwh") or 0.25),
        co2_factor_kg_per_kwh=float(data.get("co2_factor_kg_per_kwh") or 0.09),
        latitude=data.get("latitude"),
        longitude=data.get("longitude"),
        use_dynamic_weather=True,
        ventilation_enabled=True,
        ventilation_rate_ach=0.10,
    )

    db.add(cave)
    db.commit()
    db.refresh(cave)

    return {
        "status": "ok",
        "cave_id": cave.id,
        "url": f"/caves/{cave.id}",
    }